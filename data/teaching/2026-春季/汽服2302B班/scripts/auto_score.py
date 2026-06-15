#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
自动评分系统
整合抄袭检测、质量评估和提交时间数据，自动生成最终成绩
"""

import json
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed")
    print("Please install: pip install openpyxl")
    sys.exit(1)

# ==================== 配置 ====================

# 抄袭学生列表（相似度>80%）
PLAGIARISM_STUDENTS = {
    '23071140217', '23071140216', '23071140214', '23071140228',
    '23071140233', '23071140220', '23071140223', '23071140213',
    '23071140219', '23071140204', '23071140208'
}

# 提交时间（学号 -> 时间字符串）
SUBMISSION_TIMES = {
    '23071140226': '2026-06-04 16:30:28', '23071140231': '2026-06-04 16:34:35',
    '23071140230': '2026-06-06 02:28:52', '23071140217': '2026-06-06 11:04:24',
    '23071140221': '2026-06-06 12:15:50', '23071140210': '2026-06-06 12:20:02',
    '23071140233': '2026-06-06 17:00:24', '23071140235': '2026-06-06 21:41:16',
    '23071140234': '2026-06-06 21:57:37', '23071140240': '2026-06-06 22:42:25',
    '23071140241': '2026-06-06 22:48:54', '23071140238': '2026-06-06 22:56:51',
    '23071140239': '2026-06-06 23:04:58', '23071140227': '2026-06-06 23:09:14',
    '23071140201': '2026-06-07 01:05:53', '23071140229': '2026-06-07 08:33:29',
    '23071140204': '2026-06-07 08:50:07', '23071140208': '2026-06-07 08:54:01',
    '23071140237': '2026-06-07 12:21:56', '23071140213': '2026-06-07 14:09:43',
    '23071140219': '2026-06-07 14:11:53', '23071140202': '2026-06-07 14:23:46',
    '23071140214': '2026-06-07 19:20:27', '23071140228': '2026-06-07 19:36:11',
    '23071140222': '2026-06-07 19:40:57', '23071140232': '2026-06-07 19:58:16',
    '23071140215': '2026-06-07 20:45:13', '23071140206': '2026-06-07 20:53:23',
    '23071140224': '2026-06-07 21:03:28', '23071140211': '2026-06-07 22:41:34',
    '23071140220': '2026-06-07 22:52:21', '23071140203': '2026-06-07 23:27:32',
    '23071140216': '2026-06-08 09:46:04', '23071140223': '2026-06-08 09:58:02',
    '23071140236': '2026-06-08 10:15:27'
}

# 评分权重
SCORE_WEIGHTS = {
    'team_collaboration': 5,    # 团队协作
    'principle_understanding': 10,  # 实验原理与认知
    'completion': 35,          # 实验完成度（15+20）
    'code_quality': 30,        # 代码质量
    'report_quality': 10,      # 实验报告质量
    'attitude': 10             # 实验态度（教师评定）
}
TOTAL_SCORE = sum(SCORE_WEIGHTS.values())

# ==================== 自动评分逻辑 ====================

def calculate_auto_score(evaluation_data, is_plagiarism, is_submitted):
    """
    计算自动评分

    参数:
        evaluation_data: 来自evaluations.json的学生评估数据
        is_plagiarism: 是否抄袭
        is_submitted: 是否提交

    返回:
        (score, details): 分数和评分详情
    """
    if not is_submitted:
        return 0, "未提交"

    if is_plagiarism:
        return 0, "抄袭"

    # 获取质量评分
    quality_info = evaluation_data.get('quality_info', {})
    scores = evaluation_data.get('scores', {})

    # 计算各项得分
    detail_scores = {}
    total = 0

    # 1. 团队协作 (5分)
    team_score = min(scores.get('team_collaboration', 0), 5)
    detail_scores['团队协作'] = team_score
    total += team_score

    # 2. 实验原理与认知 (10分)
    principle_score = min(scores.get('principle_understanding', 0), 10)
    detail_scores['实验原理'] = principle_score
    total += principle_score

    # 3. 实验完成度 (35分)
    completion_score = min(scores.get('completion', 0), 35)
    detail_scores['实验完成度'] = completion_score
    total += completion_score

    # 4. 代码质量 (30分)
    code_score = min(scores.get('code_quality', 0), 30)
    detail_scores['代码质量'] = code_score
    total += code_score

    # 5. 实验报告质量 (10分)
    report_score = min(scores.get('report_quality', 0), 10)
    detail_scores['报告质量'] = report_score
    total += report_score

    # 6. 实验态度 (10分) - 需要教师评定，默认给满分或根据出勤
    # 这里假设全勤给满分
    attitude_score = min(scores.get('attitude', 10), 10)
    detail_scores['实验态度'] = attitude_score
    total += attitude_score

    return total, detail_scores

def get_grade(score):
    """根据分数获取等级"""
    if score >= 90:
        return 'A', '优秀'
    elif score >= 80:
        return 'B', '良好'
    elif score >= 70:
        return 'C', '中等'
    elif score >= 60:
        return 'D', '及格'
    else:
        return 'F', '不及格'

def load_evaluation_data(filepath):
    """加载评估数据"""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading evaluation data: {e}")
        return {}

# ==================== 主程序 ====================

def main():
    print("=" * 60)
    print("汽服2302B班 - 第七次实验 自动评分系统")
    print("=" * 60)

    # 加载评估数据
    eval_file = 'data/evaluations.json'
    evaluations = load_evaluation_data(eval_file)

    if not evaluations:
        print("错误: 无法加载评估数据")
        return

    # 创建学生评估映射
    eval_map = {s['student_id']: s for s in evaluations}

    # 计算所有学生的成绩
    results = []
    for sid, time_str in SUBMISSION_TIMES.items():
        is_plagiarism = sid in PLAGIARISM_STUDENTS
        is_submitted = True  # 在SUBMISSION_TIMES中的都已提交

        eval_data = eval_map.get(sid, {})

        if is_plagiarism:
            score = 0
            status = "抄袭"
            details = {"抄袭": "与相似度>80%"}
        else:
            score, details = calculate_auto_score(eval_data, False, True)
            if score == 0 and details == "未提交":
                status = "未提交"
            else:
                status = "原创"

        grade, grade_label = get_grade(score)

        results.append({
            '学号': sid,
            '提交时间': time_str,
            '总分': score,
            '等级': grade,
            '等级标签': grade_label,
            '状态': status,
            '详细得分': details
        })

    # 按提交时间排序
    results.sort(key=lambda x: x['提交时间'])

    # 生成排名
    for i, result in enumerate(results, 1):
        result['排名'] = i if result['总分'] > 0 else '-'

    # 统计信息
    total_students = len(results)
    plagiarism_count = sum(1 for r in results if r['状态'] == '抄袭')
    original_count = total_students - plagiarism_count

    avg_score = sum(r['总分'] for r in results if r['总分'] > 0) / max(original_count, 1)

    print(f"\n统计概览:")
    print(f"  总人数: {total_students}")
    print(f"  抄袭人数: {plagiarism_count} ({plagiarism_count/total_students*100:.1f}%)")
    print(f"  疑似原创: {original_count} ({original_count/total_students*100:.1f}%)")
    print(f"  平均分: {avg_score:.1f}")

    # 生成Excel报告
    generate_excel_report(results, {
        'total': total_students,
        'plagiarism': plagiarism_count,
        'original': original_count,
        'avg': avg_score
    })

    print("\n自动评分完成！")
    print("生成文件: auto_grade_report.xlsx")

def generate_excel_report(results, stats):
    """生成Excel报告"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "成绩册"

    # 标题
    ws['A1'] = '汽服2302B班 - 第七次实验 自动成绩册'
    ws.merge_cells('A1:J1')
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # 统计信息
    ws['A3'] = '统计概览'
    ws['A3'].font = Font(size=12, bold=True)

    ws['A4'] = f'总人数: {stats["total"]}'
    ws['B4'] = f'抄袭: {stats["plagiarism"]} ({stats["plagiarism"]/stats["total"]*100:.1f}%)'
    ws['C4'] = f'原创: {stats["original"]} ({stats["original"]/stats["total"]*100:.1f}%)'
    ws['D4'] = f'平均分: {stats["avg"]:.1f}'

    # 表头
    headers = ['排名', '学号', '提交时间', '总分', '等级', '等级标签', '状态',
               '团队协作', '实验原理', '实验完成度', '代码质量', '报告质量', '实验态度']

    row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # 数据行
    for result in results:
        row += 1
        details = result['详细得分']

        ws.cell(row, 1, result['排名'])
        ws.cell(row, 2, result['学号'])
        ws.cell(row, 3, result['提交时间'])
        ws.cell(row, 4, result['总分'])
        ws.cell(row, 5, result['等级'])
        ws.cell(row, 6, result['等级标签'])
        ws.cell(row, 7, result['状态'])
        ws.cell(row, 8, details.get('团队协作', 0))
        ws.cell(row, 9, details.get('实验原理', 0))
        ws.cell(row, 10, details.get('实验完成度', 0))
        ws.cell(row, 11, details.get('代码质量', 0))
        ws.cell(row, 12, details.get('报告质量', 0))
        ws.cell(row, 13, details.get('实验态度', 0))

        # 根据状态设置填充色
        if result['状态'] == '抄袭':
            fill_color = 'FFCCCC'
            for col in range(1, 14):
                ws.cell(row, col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        elif result['总分'] >= 90:
            fill_color = 'CCFFCC'
            for col in range(1, 14):
                ws.cell(row, col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

    # 设置列宽
    column_widths = [6, 12, 18, 6, 6, 8, 8, 10, 10, 12, 10, 10, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # 保存文件
    output_file = '../docs/汽服2302B班_2026春季_实验成绩册_自动评分.xlsx'
    wb.save(output_file)
    print(f"已保存: {output_file}")

if __name__ == '__main__':
    main()
