#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
更新xlsx文件：
1. 使用准确的姓名（来自答题记录）
2. 根据实验小组编号对抄袭学生分组，用颜色标注
3. 同组中最早提交的为"原创"，其他为"抄袭"
"""

import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path
import zipfile
import io
import re
from xml.etree import ElementTree as ET
from datetime import datetime

def extract_text_from_docx(docx_data):
    """从docx文件中提取文本"""
    try:
        with zipfile.ZipFile(io.BytesIO(docx_data), 'r') as docx:
            xml_content = docx.read('word/document.xml')
            root = ET.fromstring(xml_content)
            texts = []
            for elem in root.iter():
                if elem.tag.endswith('}t'):
                    if elem.text:
                        texts.append(elem.text)
            return ''.join(texts)
    except:
        return None

def get_student_info(extract_dir):
    """从答题记录和实验报告中提取学生信息"""
    student_info = {}

    for zip_file in extract_dir.glob('*.zip'):
        match = re.search(r'(\d{11})', zip_file.name)
        if not match:
            continue
        student_id = match.group(1)

        info = {'name': None, 'time': None, 'team': None}

        try:
            with zipfile.ZipFile(zip_file, 'r') as outer:
                files = outer.namelist()
                if files[0].endswith('.zip'):
                    inner_data = outer.read(files[0])
                    with zipfile.ZipFile(io.BytesIO(inner_data), 'r') as inner:
                        inner_files = inner.namelist()

                        # 从答题记录提取姓名和时间
                        doc_files = [f for f in inner_files if '答题记录' in f and f.endswith('.doc')]
                        if doc_files:
                            doc_data = inner.read(doc_files[0])
                            doc_str = str(doc_data, errors='ignore')

                            pattern1 = r'<w:t>答题人：[^<]*</w:t>\s*.*?<w:t>([^<]+)</w:t>'
                            name_match = re.search(pattern1, doc_str, re.DOTALL)
                            if name_match:
                                info['name'] = name_match.group(1).strip()

                            pattern2 = r'<w:t>提交时间：[^<]*</w:t>\s*.*?<w:t>(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})[^<]*</w:t>'
                            time_match = re.search(pattern2, doc_str, re.DOTALL)
                            if time_match:
                                info['time'] = time_match.group(1).strip()

                        # 从实验报告提取小组编号
                        docx_files = [f for f in inner_files if f.endswith('.docx') and '答题记录' not in f]
                        if docx_files:
                            docx_data = inner.read(docx_files[0])
                            text = extract_text_from_docx(docx_data)

                            if text:
                                # 搜索小组编号
                                team_match = re.search(r'第?(\d+)组|组号[：:\s]*(\d+)', text)
                                if team_match:
                                    team_num = team_match.group(1) or team_match.group(2)
                                    info['team'] = team_num

                        if any(info.values()):
                            student_info[student_id] = info

        except Exception as e:
            pass

    return student_info

def update_xlsx_with_groups(xlsx_file, extract_dir):
    """更新xlsx文件"""
    print(f"处理文件: {xlsx_file.name}")

    # 获取学生信息
    print("从答题记录和实验报告提取信息...")
    student_info = get_student_info(extract_dir)
    print(f"  提取到 {len(student_info)} 个学生信息")

    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb['实验7-档位模拟器']

    # 清除现有颜色
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.fill = PatternFill(fill_type=None)

    # 定义颜色
    colors = [
        'FFFF00',  # 黄色
        '00FFFF',  # 青色
        'FF00FF',  # 紫色
        'FF9900',  # 橙色
        '00FF00',  # 绿色
        'FF0000',  # 红色
        'CCFFCC',  # 浅绿
        'FFCC99',  # 浅橙
    ]

    # 找到列位置
    name_col = time_col = id_col = status_col = team_col = None
    header_row = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5), 1):
        for col_idx, cell in enumerate(row, 1):
            val = cell.value
            if val == '姓名':
                name_col = col_idx
            elif val == '提交时间':
                time_col = col_idx
            elif val == '学号':
                id_col = col_idx
            elif val == '状态':
                status_col = col_idx
        if name_col and id_col:
            header_row = row_idx
            break

    # 1. 更新姓名和提交时间
    print("\n更新姓名和提交时间...")
    updated = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        id_cell = ws.cell(row_idx, id_col)
        student_id = str(id_cell.value).strip() if id_cell.value else ''

        if student_id and student_id in student_info:
            info = student_info[student_id]

            if name_col and info['name']:
                ws.cell(row_idx, name_col).value = info['name']
                updated += 1

            if time_col and info['time']:
                ws.cell(row_idx, time_col).value = info['time']

    print(f"  更新了 {updated} 个姓名")

    # 2. 分析抄袭学生并按实验小组分组
    print("\n分析抄袭学生...")

    # 收集抄袭学生信息
    plagiarism_students = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        id_cell = ws.cell(row_idx, id_col)
        status_cell = ws.cell(row_idx, status_col)

        if status_cell.value == '抄袭':
            student_id = str(id_cell.value).strip() if id_cell.value else ''
            if student_id and student_id in student_info:
                info = student_info[student_id]
                if info['time']:
                    plagiarism_students.append({
                        'id': student_id,
                        'name': info['name'],
                        'time': info['time'],
                        'team': info['team'],
                        'row': row_idx
                    })

    # 按实验小组分组
    team_groups = {}
    for s in plagiarism_students:
        team = s['team'] or '未知'
        if team not in team_groups:
            team_groups[team] = []
        team_groups[team].append(s)

    print(f"  找到 {len(team_groups)} 个抄袭小组:")

    # 对每组按提交时间排序，找出原创者
    color_idx = 0
    for team, students in sorted(team_groups.items()):
        color = colors[color_idx % len(colors)]
        students.sort(key=lambda x: x['time'])  # 按时间排序

        print(f"\n  小组 {team} ({len(students)}人) - 颜色: {color}")

        for i, s in enumerate(students):
            role = "原创" if i == 0 else "抄袭"
            print(f"    {s['time']} - {s['id']} - {s['name']} - {role}")

            # 标注整行
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(s['row'], col_idx)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

            # 更新状态：第一个是原创，其他是抄袭
            ws.cell(s['row'], status_col).value = role

        color_idx += 1

    # 保存
    wb.save(xlsx_file)
    wb.close()

    print(f"\n文件已保存: {xlsx_file.name}")
    print(f"  共标注 {len(team_groups)} 组抄袭学生")


def main():
    xlsx_file = Path(__file__).parent.parent / 'docs/teaching/2026-春季/汽服2302B班/docs/汽服2302B班_2026春季学期成绩册.xlsx'
    extract_dir = Path(__file__).parent.parent / 'docs/teaching/2026-春季/汽服2302B班/07-car-gear/submissions/extracted'

    if not xlsx_file.exists():
        print(f"文件不存在: {xlsx_file}")
        return

    if not extract_dir.exists():
        print(f"目录不存在: {extract_dir}")
        return

    update_xlsx_with_groups(xlsx_file, extract_dir)


if __name__ == '__main__':
    main()
