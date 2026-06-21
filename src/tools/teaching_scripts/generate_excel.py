#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成抄袭检测分析Excel报告"""

from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

# 产物统一写入仓库根目录下的 outputs/reports，避免写到进程 CWD 污染源码树
_OUTPUT_DIR = Path(__file__).resolve().parents[3] / "outputs" / "reports"


def main():
    if not _OPENPYXL_OK:
        print('Error: openpyxl 未安装')
        print('Please install openpyxl: pip install openpyxl')
        return

    times = {
        '23071140226': '2026-06-04 16:30:28', '23071140231': '2026-06-04 16:34:35',
        '23071140230': '2026-06-06 02:28:52', '23071140217': '2026-06-06 11:04:24',
        '23071140221': '2026-06-06 12:15:50', '23071140210': '2026-06-06 12:20:02',
        '23071140233': '2026-06-06 17:00:24', '23071140235': '2026-06-06 21:41:16',
        '23071140234': '2026-06-06 21:57:37', '23071140240': '2026-06-06 22:42:25',
        '23071140241': '2026-06-06 22:48:54', '23071140238': '2026-06-06 22:56:51',
        '23071140239': '2026-06-06 23:04:58', '23071140227': '2026-06-06 23:09:14',
        '23071140201': '2026-06-07 01:05:53', '23071140229': '2026-06-07 08:33:29',
        '23071140204': '2026-06-07 08:50:07', '23071140208': '2026-06-07 08:54:01',
        '23071140237': '2026-06-07 12:21:56', '23071140213': '2026-06-07 14:09:43',
        '23071140219': '2026-06-07 14:11:53', '23071140202': '2026-06-07 14:23:46',
        '23071140214': '2026-06-07 19:20:27', '23071140228': '2026-06-07 19:36:11',
        '23071140222': '2026-06-07 19:40:57', '23071140232': '2026-06-07 19:58:16',
        '23071140215': '2026-06-07 20:45:13', '23071140206': '2026-06-07 20:53:23',
        '23071140224': '2026-06-07 21:03:28', '23071140211': '2026-06-07 22:41:34',
        '23071140220': '2026-06-07 22:52:21', '23071140203': '2026-06-07 23:27:32',
        '23071140216': '2026-06-08 09:46:04', '23071140223': '2026-06-08 09:58:02',
        '23071140236': '2026-06-08 10:15:27'
    }

    # Plagiarism pairs with similarity details
    plagiarism_pairs = {
        '23071140217': {'similar_to': '23071140216', 'similarity': '100%', 'note': '完全相同'},
        '23071140216': {'similar_to': '23071140217,23071140223', 'similarity': '100%', 'note': '与两人完全相同'},
        '23071140214': {'similar_to': '23071140228', 'similarity': '100%', 'note': '完全相同'},
        '23071140228': {'similar_to': '23071140214', 'similarity': '100%', 'note': '完全相同'},
        '23071140233': {'similar_to': '23071140220', 'similarity': '96.6%', 'note': '心得总结高度相似'},
        '23071140220': {'similar_to': '23071140233', 'similarity': '96.6%', 'note': '心得总结高度相似'},
        '23071140223': {'similar_to': '23071140216', 'similarity': '100%', 'note': '完全相同'},
        '23071140213': {'similar_to': '23071140219', 'similarity': '95.9%', 'note': '心得总结高度相似'},
        '23071140219': {'similar_to': '23071140213', 'similarity': '95.9%', 'note': '心得总结高度相似'},
        '23071140204': {'similar_to': '23071140208', 'similarity': '96.3%', 'note': '心得总结高度相似'},
        '23071140208': {'similar_to': '23071140204', 'similarity': '96.3%', 'note': '心得总结高度相似'},
    }

    zero_points = list(plagiarism_pairs.keys())

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ============ Sheet 1: Summary ============
    ws1 = wb.create_sheet('汇总')

    ws1['A1'] = '汽服2302B班 - 第七次实验 抄袭检测汇总报告'
    ws1.merge_cells('A1:D1')
    ws1['A1'].font = Font(size=16, bold=True)
    ws1['A1'].alignment = Alignment(horizontal='center')

    ws1['A3'] = '统计概览'
    ws1['A3'].font = Font(size=14, bold=True)

    ws1['A5'] = '总人数'
    ws1['B5'] = len(times)
    ws1['B5'].font = Font(bold=True)

    ws1['A6'] = '抄袭人数'
    ws1['B6'] = len(zero_points)
    ws1['B6'].font = Font(color='FF0000', bold=True)

    ws1['A7'] = '疑似原创人数'
    ws1['B7'] = len(times) - len(zero_points)
    ws1['B7'].font = Font(color='00AA00', bold=True)

    ws1['A9'] = '抄袭率'
    rate = f'{len(zero_points)/len(times)*100:.1f}%'
    ws1['B9'] = rate
    ws1['B9'].font = Font(color='FF0000', bold=True)

    ws1['A11'] = '说明'
    ws1['A11'].font = Font(bold=True)
    ws1['A12'] = '1. 抄袭者与被抄袭者均记0分'
    ws1['A13'] = '2. 心得总结相似度>80%判定为抄袭'
    ws1['A14'] = '3. 疑似原创学生需人工复核确认'

    for col in range(1, 5):
        ws1.column_dimensions[get_column_letter(col)].width = 20

    # ============ Sheet 2: Detailed Scoring ============
    ws2 = wb.create_sheet('详细评分')

    ws2['A1'] = '学生详细评分表'
    ws2.merge_cells('A1:F1')
    ws2['A1'].font = Font(size=14, bold=True)
    ws2['A1'].alignment = Alignment(horizontal='center')

    headers = ['排名', '学号', '提交时间', '评分', '状态', '相似对象']
    for i, h in enumerate(headers, 1):
        cell = ws2.cell(3, i, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

    all_students = sorted(times.keys(), key=lambda x: times[x])
    row = 4
    for rank, sid in enumerate(all_students, 1):
        time = times[sid]

        if sid in zero_points:
            score = '0分'
            status = '抄袭'
            similar = plagiarism_pairs[sid]['similar_to']
            fill_color = 'FFCCCC'
        else:
            score = '待定'
            status = '疑似原创'
            similar = '-'
            fill_color = None

        ws2.cell(row, 1, rank)
        ws2.cell(row, 2, sid)
        ws2.cell(row, 3, time)
        ws2.cell(row, 4, score)
        ws2.cell(row, 5, status)
        ws2.cell(row, 6, similar)

        if fill_color:
            for col in range(1, 7):
                ws2.cell(row, col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        row += 1

    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # ============ Sheet 3: Plagiarism Pairs ============
    ws3 = wb.create_sheet('抄袭对')

    ws3['A1'] = '抄袭关系详细分析'
    ws3.merge_cells('A1:G1')
    ws3['A1'].font = Font(size=14, bold=True)
    ws3['A1'].alignment = Alignment(horizontal='center')

    headers = ['群组', '学号1', '提交时间1', '学号2', '提交时间2', '相似度', '说明']
    for i, h in enumerate(headers, 1):
        cell = ws3.cell(3, i, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

    pairs = [
        ('第1组', '23071140217', times['23071140217'], '23071140216', times['23071140216'], '100%', '报告完全相同'),
        ('第1组', '23071140216', times['23071140216'], '23071140223', times['23071140223'], '100%', '报告完全相同'),
        ('第2组', '23071140214', times['23071140214'], '23071140228', times['23071140228'], '100%', '报告完全相同'),
        ('第3组', '23071140233', times['23071140233'], '23071140220', times['23071140220'], '96.6%', '心得总结高度相似'),
        ('第4组', '23071140213', times['23071140213'], '23071140219', times['23071140219'], '95.9%', '心得总结高度相似'),
        ('第5组', '23071140204', times['23071140204'], '23071140208', times['23071140208'], '96.3%', '心得总结高度相似'),
    ]

    row = 4
    for pair in pairs:
        for col, val in enumerate(pair, 1):
            ws3.cell(row, col, val)
        row += 1

    for col in range(1, 8):
        ws3.column_dimensions[get_column_letter(col)].width = 16

    # ============ Sheet 4: Zero Points List ============
    ws4 = wb.create_sheet('0分名单')

    ws4['A1'] = '记0分学生详细名单'
    ws4.merge_cells('A1:E1')
    ws4['A1'].font = Font(size=14, bold=True)
    ws4['A1'].alignment = Alignment(horizontal='center')

    headers = ['学号', '提交时间', '相似对象', '相似度', '抄袭说明']
    for i, h in enumerate(headers, 1):
        cell = ws4.cell(3, i, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

    row = 4
    for sid in sorted(zero_points):
        info = plagiarism_pairs[sid]
        ws4.cell(row, 1, sid)
        ws4.cell(row, 2, times[sid])
        ws4.cell(row, 3, info['similar_to'])
        ws4.cell(row, 4, info['similarity'])
        ws4.cell(row, 5, info['note'])
        row += 1

    for col in range(1, 6):
        ws4.column_dimensions[get_column_letter(col)].width = 20

    # ============ Sheet 5: Original Students ============
    ws5 = wb.create_sheet('疑似原创')

    ws5['A1'] = '疑似原创学生名单（需人工复核）'
    ws5.merge_cells('A1:C1')
    ws5['A1'].font = Font(size=14, bold=True)
    ws5['A1'].alignment = Alignment(horizontal='center')

    ws5['A3'] = '排名'
    ws5['B3'] = '学号'
    ws5['C3'] = '提交时间'
    for i in range(1, 4):
        cell = ws5.cell(3, i)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

    row = 4
    rank = 1
    for sid in sorted(times.keys(), key=lambda x: times[x]):
        if sid not in zero_points:
            ws5.cell(row, 1, rank)
            ws5.cell(row, 2, sid)
            ws5.cell(row, 3, times[sid])
            row += 1
            rank += 1

    ws5.cell(row + 1, 1, '合计')
    ws5.cell(row + 1, 2, rank - 1)
    ws5.cell(row + 1, 2).font = Font(bold=True)

    for col in range(1, 4):
        ws5.column_dimensions[get_column_letter(col)].width = 18

    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = _OUTPUT_DIR / 'plagiarism_analysis_report.xlsx'
    wb.save(out_path)
    print(f'Excel report generated successfully: {out_path}')
    print(f'Contains 5 sheets: 汇总, 详细评分, 抄袭对, 0分名单, 疑似原创')


if __name__ == '__main__':
    main()
