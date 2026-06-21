#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
更新成绩册脚本
用新的评分结果更新Excel成绩册
"""

import json
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# 配置路径
SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR / 'data'
EXCEL_DIR = SCRIPT_DIR.parent / 'docs'

# 成绩册文件
GRADEBOOK_PATH = EXCEL_DIR / '汽服2302B班_2026春季学期成绩册_updated_20260610_010822.xlsx'

# 评分数据文件
EVALUATIONS_PATH = DATA_DIR / 'evaluations.json'
STUDENTS_PATH = DATA_DIR / 'students.json'
QUALITY_PATH = DATA_DIR / 'quality_assessment.json'
EXTRACTED_PATH = DATA_DIR / 'extracted_content.json'


def load_extracted_content():
    """加载提取的内容数据"""
    with open(EXTRACTED_PATH, 'r', encoding='utf-8') as f:
        content_list = json.load(f)

    # 转换为字典
    content_map = {}
    for content in content_list:
        student_id = content.get('student_id', '')
        content_map[student_id] = content

    return content_map


def load_evaluations():
    """加载评分数据"""
    with open(EVALUATIONS_PATH, 'r', encoding='utf-8') as f:
        evaluations = json.load(f)

    # 创建学号到评分的映射
    eval_map = {}
    for eval_data in evaluations:
        eval_map[eval_data['student_id']] = eval_data

    return eval_map


def load_students():
    """加载学生数据"""
    with open(STUDENTS_PATH, 'r', encoding='utf-8') as f:
        students_list = json.load(f)

    # 转换为字典，以学号为键
    students_dict = {}
    for student in students_list:
        student_id = student.get('id', '')
        students_dict[student_id] = student

    return students_dict


def load_quality_data():
    """加载质量评估数据"""
    if not QUALITY_PATH.exists():
        return {}

    with open(QUALITY_PATH, 'r', encoding='utf-8') as f:
        data = json.load(f)
        return data.get('quality_scores', {})


def calculate_similarity(eval_data, quality_data):
    """获取抄袭相似度"""
    student_id = eval_data['student_id']

    if student_id in quality_data:
        quality_info = quality_data[student_id]
        # 获取最高相似度
        max_sim = 0
        if quality_info.get('similar_pairs'):
            for pair in quality_info['similar_pairs']:
                sim = pair.get('overall', 0)
                if sim > max_sim:
                    max_sim = sim
        return max_sim if max_sim > 0 else 0  # 0-100 百分比（与 overall 原始单位、determine_status 阈值、plagiarism_status 分支一致）

    return 0


def determine_status(eval_data, similarity):
    """确定学生状态"""
    total_score = eval_data.get('total_score', 0)

    # 高相似度(>90%) = 抄袭（严格判定）
    if similarity > 90:
        return '抄袭'

    # 中高相似度(60-90%) = 警告
    if similarity > 60:
        return '警告'

    # 0分 = 未提交
    if total_score == 0:
        return '未提交'

    return '原创'


def update_gradebook():
    """更新成绩册"""
    print("=== 更新成绩册 ===\n")

    # 加载数据
    print("1. 加载评分数据...")
    evaluations = load_evaluations()
    students = load_students()
    quality_data = load_quality_data()
    extracted_data = load_extracted_content()

    print(f"   - 评分记录: {len(evaluations)} 条")
    print(f"   - 学生记录: {len(students)} 条")
    print(f"   - 质量评估: {len(quality_data)} 条")
    print(f"   - 提取内容: {len(extracted_data)} 条")

    # 检查是否有抄袭惩罚数据
    has_plagiarism_data = any('plagiarism_status' in e for e in evaluations)
    if has_plagiarism_data:
        print(f"   - 检测到抄袭惩罚数据")

    # 加载Excel工作簿
    print("\n2. 加载Excel成绩册...")
    wb = load_workbook(GRADEBOOK_PATH)
    print(f"   - 工作表: {wb.sheetnames}")

    # 更新实验7工作表
    print("\n3. 更新实验7成绩...")

    # 找到实验7工作表
    exp7_sheet = None
    for sheet_name in wb.sheetnames:
        if '实验7' in sheet_name or '档位' in sheet_name:
            exp7_sheet = wb[sheet_name]
            break

    if not exp7_sheet:
        print("   错误: 找不到实验7工作表")
        return

    # 找到数据行范围（从第3行开始）
    print(f"   - 工作表: {exp7_sheet.title}")

    # 创建新的数据列表
    updated_rows = []
    stats = {
        'total': 0,
        'original': 0,
        'plagiarism': 0,
        'no_submit': 0,
        'warning': 0,
        'total_score': 0
    }

    # 按学号排序学生
    sorted_student_ids = sorted(evaluations.keys(), key=lambda x: int(x) if x.isdigit() else 0)

    for idx, student_id in enumerate(sorted_student_ids, 1):
        eval_data = evaluations[student_id]
        student_info = students.get(student_id, {})
        extracted_info = extracted_data.get(student_id, {})

        # 确定状态 - 优先使用已计算的抄袭状态
        if 'plagiarism_status' in eval_data:
            status = eval_data['plagiarism_status']
            similarity = eval_data.get('similarity', 0)
        else:
            similarity = calculate_similarity(eval_data, quality_data)
            status = determine_status(eval_data, similarity)

        # 更新统计
        stats['total'] += 1
        if status == '原创':
            stats['original'] += 1
            stats['total_score'] += eval_data.get('total_score', 0)
        elif status == '抄袭':
            stats['plagiarism'] += 1
        elif status == '未提交':
            stats['no_submit'] += 1
        elif status == '警告':
            stats['warning'] += 1
            stats['original'] += 1
            stats['total_score'] += eval_data.get('total_score', 0)

        # 获取分数
        scores = eval_data.get('scores', {})
        total_score = eval_data.get('total_score', 0)
        grade = eval_data.get('grade', 'F')

        # 获取学生信息
        name = eval_data.get('name', student_info.get('name', ''))

        # 从提取的数据中获取其他信息
        analysis = extracted_info.get('analysis', {})
        team_members = analysis.get('team_members', [])
        group = team_members[0] if team_members else ''

        # 从docx_path中提取提交时间（文件修改时间）
        docx_path = extracted_info.get('docx_path', '')
        submit_time = ''
        if docx_path and Path(docx_path).exists():
            try:
                timestamp = Path(docx_path).stat().st_mtime
                submit_time = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
            except (OSError, ValueError, OverflowError):
                # 文件不可访问或时间戳非法，留空
                pass

        # 准备行数据
        row_data = [
            idx,                           # 序号
            student_id,                    # 学号
            name,                          # 姓名
            group,                         # 实验小组
            f"{similarity:.1f}%",          # 相似度（0-100）
            submit_time,                   # 提交时间
            total_score,                   # 总分
            grade,                         # 等级
            status,                        # 状态
            scores.get('team_collaboration', 0),   # 团队协作
            scores.get('principle_understanding', 0), # 实验原理
            scores.get('completion', 0),            # 实验完成度
            scores.get('code_quality', 0),          # 代码质量
            scores.get('report_quality', 0),        # 报告质量
        ]

        updated_rows.append(row_data)

    # 计算平均分
    avg_score = stats['total_score'] / stats['original'] if stats['original'] > 0 else 0

    # 清空现有数据（从第4行开始）
    print(f"   - 清空旧数据...")
    for row in range(4, exp7_sheet.max_row + 1):
        for col in range(1, exp7_sheet.max_column + 1):
            cell = exp7_sheet.cell(row=row, column=col)
            if row != 4:  # 保留第4行（表头）
                cell.value = None

    # 写入新数据
    print(f"   - 写入新数据...")
    start_row = 5  # 从第5行开始写数据

    for row_idx, row_data in enumerate(updated_rows, start_row):
        for col_idx, value in enumerate(row_data, 1):
            cell = exp7_sheet.cell(row=row_idx, column=col_idx, value=value)

            # 设置对齐
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # 根据状态设置颜色
            if col_idx == 9:  # 状态列
                if value == '抄袭':
                    cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    cell.font = Font(color='FFFFFF', bold=True)
                elif value == '未提交':
                    cell.fill = PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid')
                    cell.font = Font(bold=True)
                elif value == '警告':
                    cell.fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                    cell.font = Font(bold=True)

            # 根据等级设置颜色
            if col_idx == 8:  # 等级列
                if value == 'A':
                    cell.font = Font(color='008000', bold=True)
                elif value == 'F':
                    cell.font = Font(color='FF0000', bold=True)

    # 更新统计行（第1行）
    print(f"   - 更新统计信息...")
    stats_cell = exp7_sheet['A1']
    stats_text = f"统计: 共{stats['total']}人 | 原创{stats['original']}人 | 抄袭{stats['plagiarism']}人 | 未提交{stats['no_submit']}人 | 平均{avg_score:.1f}分"
    stats_cell.value = stats_text

    # 更新学期汇总工作表
    print("\n4. 更新学期汇总...")
    summary_sheet = wb['学期汇总']

    # 找到实验7列并更新
    header_row = 2
    for col in range(1, summary_sheet.max_column + 1):
        cell = summary_sheet.cell(row=header_row, column=col)
        if cell.value and '实验7' in str(cell.value):
            # 更新这一列的成绩
            for row in range(3, summary_sheet.max_row + 1):
                student_id_cell = summary_sheet.cell(row=row, column=1)
                if student_id_cell.value:
                    student_id = str(student_id_cell.value)
                    if student_id in evaluations:
                        eval_data = evaluations[student_id]
                        score_cell = summary_sheet.cell(row=row, column=col)
                        score_cell.value = eval_data.get('total_score', 0)

                        # 设置格式
                        score_cell.alignment = Alignment(horizontal='center')
                        score = eval_data.get('total_score', 0)
                        if score >= 90:
                            score_cell.font = Font(color='008000')
                        elif score < 60:
                            score_cell.font = Font(color='FF0000')
            break

    # 保存文件
    print("\n5. 保存更新后的成绩册...")

    # 生成新文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    new_filename = f'汽服2302B班_2026春季学期成绩册_updated_{timestamp}.xlsx'
    new_path = EXCEL_DIR / new_filename

    wb.save(new_path)

    print(f"\n✅ 成绩册更新完成!")
    print(f"   新文件: {new_path}")
    print(f"\n=== 统计信息 ===")
    print(f"   总人数: {stats['total']}")
    print(f"   原创: {stats['original']} 人")
    print(f"   抄袭: {stats['plagiarism']} 人")
    print(f"   未提交: {stats['no_submit']} 人")
    print(f"   警告: {stats['warning']} 人")
    print(f"   平均分: {avg_score:.1f}")


if __name__ == '__main__':
    try:
        update_gradebook()
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()
