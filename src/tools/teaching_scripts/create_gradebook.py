#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""创建班级实验成绩册Excel文件"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 学生提交时间数据
times = {
    '23071140226': '2026-06-04 16:30:28', '23071140231': '2026-06-04 16:34:35',
    '23071140230': '2026-06-06 02:28:52', '23071140217': '2026-06-06 11:04:24',
    '23071140221': '2026-06-06 12:15:50', '23071140210': '2026-06-06 12:20:02',
    '23071140233': '2026-06-06 17:00:24', '23071140235': '2026-06-06 21:41:16',
    '23071140234': '2026-06-06 21:57:37', '23071140240': '2026-06-06 22:42:25',
    '23071140241': '2026-06-06 22:48:54', '23071140238': '2026-06-06 22:56:51',
    '23071140239': '2026-06-06 23:04:58', '23071140227': '2026-06-06 23:09:14',
    '23071140201': '2026-06-07 01:05:53', '23071140229': '2026-06-07 08:33:29',
    '23071140204': '2026-06-07 08:50:07', '23071140208': '2026-06-07 08:54:01',
    '23071140237': '2026-06-07 12:21:56', '23071140213': '2026-06-07 14:09:43',
    '23071140219': '2026-06-07 14:11:53', '23071140202': '2026-06-07 14:23:46',
    '23071140214': '2026-06-07 19:20:27', '23071140228': '2026-06-07 19:36:11',
    '23071140222': '2026-06-07 19:40:57', '23071140232': '2026-06-07 19:58:16',
    '23071140215': '2026-06-07 20:45:13', '23071140206': '2026-06-07 20:53:23',
    '23071140224': '2026-06-07 21:03:28', '23071140211': '2026-06-07 22:41:34',
    '23071140220': '2026-06-07 22:52:21', '23071140203': '2026-06-07 23:27:32',
    '23071140216': '2026-06-08 09:46:04', '23071140223': '2026-06-08 09:58:02',
    '23071140236': '2026-06-08 10:15:27'
}

# 抄袭对数据
plagiarism_pairs = {
    '23071140217': {'similar_to': '23071140216', 'similarity': '100%'},
    '23071140216': {'similar_to': '23071140217,23071140223', 'similarity': '100%'},
    '23071140214': {'similar_to': '23071140228', 'similarity': '100%'},
    '23071140228': {'similar_to': '23071140214', 'similarity': '100%'},
    '23071140233': {'similar_to': '23071140220', 'similarity': '96.6%'},
    '23071140220': {'similar_to': '23071140233', 'similarity': '96.6%'},
    '23071140223': {'similar_to': '23071140216', 'similarity': '100%'},
    '23071140213': {'similar_to': '23071140219', 'similarity': '95.9%'},
    '23071140219': {'similar_to': '23071140213', 'similarity': '95.9%'},
    '23071140204': {'similar_to': '23071140208', 'similarity': '96.3%'},
    '23071140208': {'similar_to': '23071140204', 'similarity': '96.3%'},
}

zero_points = list(plagiarism_pairs.keys())

# 创建工作簿
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============ Sheet 1: 成绩单 ============
ws1 = wb.create_sheet('成绩单')

ws1['A1'] = '汽服2302B班 - 2026春季学期 实验成绩单'
ws1.merge_cells('A1:K1')
ws1['A1'].font = Font(size=16, bold=True)
ws1['A1'].alignment = Alignment(horizontal='center')

ws1['A2'] = '学期: 2026春季 | 班级: 汽服2302B班'
ws1.merge_cells('A2:K2')
ws1['A2'].alignment = Alignment(horizontal='center')

# 表头
headers = ['学号', '姓名', '实验1', '实验2', '实验3', '实验4', '实验5', '实验6', '实验7-档位', '实验8', '平均分']
for i, h in enumerate(headers, 1):
    cell = ws1.cell(4, i, h)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# 学生列表（按学号排序）
all_students = sorted(times.keys())
row = 5
for sid in all_students:
    ws1.cell(row, 1, sid)
    ws1.cell(row, 2, '')  # 姓名预留

    # 实验7分数
    if sid in zero_points:
        ws1.cell(row, 9, 0)
        ws1.cell(row, 9).font = Font(color='FF0000', bold=True)
        # 整行标记
        for col in range(1, 12):
            ws1.cell(row, col).fill = PatternFill(start_color='FFEEEE', end_color='FFEEEE', fill_type='solid')
    else:
        ws1.cell(row, 9, '待定')

    row += 1

# 统计行
ws1.cell(row, 1, '统计')
ws1.cell(row, 1).font = Font(bold=True)
ws1.cell(row, 9, f'0分: {len(zero_points)}人')
ws1.cell(row, 9).font = Font(color='FF0000', bold=True)

for col in range(1, 12):
    ws1.column_dimensions[get_column_letter(col)].width = 12

# ============ Sheet 2: 实验7详情 ============
ws2 = wb.create_sheet('实验7-档位')

ws2['A1'] = '实验7: 汽车档位模拟器设计 - 详细分析'
ws2.merge_cells('A1:H1')
ws2['A1'].font = Font(size=14, bold=True)
ws2['A1'].alignment = Alignment(horizontal='center')

ws2['A2'] = '实验日期: 2026年6月 | 提交人数: 35人 | 抄袭: 11人'
ws2.merge_cells('A2:H2')
ws2['A2'].alignment = Alignment(horizontal='center')

headers = ['学号', '提交时间', '评分', '状态', '相似对象', '相似度', '说明', '反馈文件']
for i, h in enumerate(headers, 1):
    cell = ws2.cell(4, i, h)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

all_students_sorted = sorted(times.keys(), key=lambda x: times[x])
row = 5
for sid in all_students_sorted:
    time = times[sid]

    if sid in zero_points:
        score = '0'
        status = '抄袭'
        similar = plagiarism_pairs[sid]['similar_to']
        sim = plagiarism_pairs[sid]['similarity']
        note = '心得总结高度相似'
        fill_color = 'FFCCCC'
    else:
        score = '待定'
        status = '疑似原创'
        similar = '-'
        sim = '-'
        note = '需人工复核'
        fill_color = None

    ws2.cell(row, 1, sid)
    ws2.cell(row, 2, time)
    ws2.cell(row, 3, score)
    ws2.cell(row, 4, status)
    ws2.cell(row, 5, similar)
    ws2.cell(row, 6, sim)
    ws2.cell(row, 7, note)
    ws2.cell(row, 8, f'{sid}_反馈.docx')

    if fill_color:
        for col in range(1, 9):
            ws2.cell(row, col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    row += 1

for col in range(1, 9):
    ws2.column_dimensions[get_column_letter(col)].width = 16

# ============ Sheet 3: 实验7抄袭分析 ============
ws3 = wb.create_sheet('实验7-抄袭')

ws3['A1'] = '实验7: 抄袭关系详细分析'
ws3.merge_cells('A1:G1')
ws3['A1'].font = Font(size=14, bold=True)
ws3['A1'].alignment = Alignment(horizontal='center')

headers = ['群组', '学号1', '提交时间1', '学号2', '提交时间2', '相似度', '说明']
for i, h in enumerate(headers, 1):
    cell = ws3.cell(3, i, h)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

pairs = [
    ('G1', '23071140217', times['23071140217'], '23071140216', times['23071140216'], '100%', '完全相同'),
    ('G1', '23071140216', times['23071140216'], '23071140223', times['23071140223'], '100%', '完全相同'),
    ('G2', '23071140214', times['23071140214'], '23071140228', times['23071140228'], '100%', '完全相同'),
    ('G3', '23071140233', times['23071140233'], '23071140220', times['23071140220'], '96.6%', '高度相似'),
    ('G4', '23071140213', times['23071140213'], '23071140219', times['23071140219'], '95.9%', '高度相似'),
    ('G5', '23071140204', times['23071140204'], '23071140208', times['23071140208'], '96.3%', '高度相似'),
]

row = 4
for pair in pairs:
    for col, val in enumerate(pair, 1):
        ws3.cell(row, col, val)
    ws3.cell(row, 7).fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    row += 1

for col in range(1, 8):
    ws3.column_dimensions[get_column_letter(col)].width = 16

# ============ Sheet 4: 实验7统计 ============
ws4 = wb.create_sheet('实验7-统计')

ws4['A1'] = '实验7: 统计概览'
ws4.merge_cells('A1:C1')
ws4['A1'].font = Font(size=14, bold=True)

ws4['A3'] = '统计项'
ws4['B3'] = '数值'
ws4['C3'] = '占比'
for col in range(1, 4):
    ws4.cell(3, col).font = Font(bold=True)

stats = [
    ('总提交人数', len(times), f'{len(times)}人'),
    ('抄袭人数', len(zero_points), f'{len(zero_points)/len(times)*100:.1f}%'),
    ('疑似原创人数', len(times)-len(zero_points), f'{(len(times)-len(zero_points))/len(times)*100:.1f}%'),
]

row = 4
for stat in stats:
    ws4.cell(row, 1, stat[0])
    ws4.cell(row, 2, stat[1])
    ws4.cell(row, 3, stat[2])
    if '抄袭' in stat[0]:
        ws4.cell(row, 2).font = Font(color='FF0000', bold=True)
    row += 1

for col in range(1, 4):
    ws4.column_dimensions[get_column_letter(col)].width = 18

# 保存文件
filename = '汽服2302B班_2026春季_实验成绩册.xlsx'
wb.save(filename)
print(f'Generated: {filename}')
print('Sheets: 成绩单, 实验7-档位, 实验7-抄袭, 实验7-统计')
