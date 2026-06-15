#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
合并班级xlsx文件
将多个xlsx文件的独特sheet合并到一个文件中
"""

import os
import shutil
from pathlib import Path
import openpyxl

def merge_xlsx_files(base_dir):
    """合并xlsx文件"""
    base_dir = Path(base_dir)
    xlsx_files = list(base_dir.glob('*.xlsx'))

    if len(xlsx_files) <= 1:
        print(f"只有一个或没有xlsx文件，无需合并")
        return

    print(f"找到 {len(xlsx_files)} 个xlsx文件:")
    for f in xlsx_files:
        print(f"  - {f.name}")

    # 选择要保留的目标文件（学期成绩册）
    target_name = '汽服2302B班_2026春季学期成绩册.xlsx'
    target_file = None

    for f in xlsx_files:
        if f.name == target_name:
            target_file = f
            break

    if not target_file:
        # 如果找不到目标文件，使用第一个文件
        target_file = xlsx_files[0]
        target_name = target_file.name

    print(f"\n目标文件: {target_name}")
    print(f"\n正在合并...\n")

    # 打开目标文件
    target_wb = openpyxl.load_workbook(target_file)
    existing_sheets = set(target_wb.sheetnames)

    print(f"目标文件现有sheet: {existing_sheets}")

    # 遍历其他文件，复制独特的sheet
    files_to_delete = []

    for source_file in xlsx_files:
        if source_file == target_file:
            continue

        print(f"\n处理: {source_file.name}")
        try:
            source_wb = openpyxl.load_workbook(source_file)

            for sheet_name in source_wb.sheetnames:
                # 如果sheet已存在，跳过
                if sheet_name in existing_sheets:
                    print(f"  跳过已存在的sheet: {sheet_name}")
                    continue

                # 复制sheet
                print(f"  复制sheet: {sheet_name}")
                source_ws = source_wb[sheet_name]
                target_ws = target_wb.create_sheet(sheet_name)

                # 复制数据（简化版，不复制样式以避免错误）
                max_row = source_ws.max_row
                max_col = source_ws.max_column

                for row_idx in range(1, max_row + 1):
                    for col_idx in range(1, max_col + 1):
                        source_cell = source_ws.cell(row_idx, col_idx)
                        target_cell = target_ws.cell(row_idx, col_idx)
                        target_cell.value = source_cell.value

                existing_sheets.add(sheet_name)

            source_wb.close()
            files_to_delete.append(source_file)

        except Exception as e:
            print(f"  错误: {e}")

    # 保存合并后的文件
    backup_name = target_file.stem + '_backup' + target_file.suffix
    backup_path = base_dir / backup_name

    print(f"\n备份原文件到: {backup_name}")
    shutil.copy2(target_file, backup_path)

    print(f"保存合并后的文件: {target_file}")
    target_wb.save(target_file)
    target_wb.close()

    print(f"\n合并完成！目标文件现在包含 {len(target_wb.sheetnames)} 个sheet")

    # 删除其他文件
    print(f"\n删除以下文件:")
    for f in files_to_delete:
        print(f"  删除: {f.name}")
        f.unlink()

    print(f"\n最终保留: {target_name}")

    return target_file


def main():
    base_dir = Path(__file__).parent.parent / 'docs/teaching/2026-春季/汽服2302B班/docs'

    if not base_dir.exists():
        print(f"目录不存在: {base_dir}")
        return

    merge_xlsx_files(base_dir)


if __name__ == '__main__':
    main()
