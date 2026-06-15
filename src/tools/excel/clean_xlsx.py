#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
清理xlsx文件：
1. 用颜色标注相似度高（抄袭）的学生
2. 删除多余的sheet
"""

import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path

def clean_xlsx(file_path):
    """清理xlsx文件"""
    file_path = Path(file_path)

    print(f"处理文件: {file_path.name}")

    wb = openpyxl.load_workbook(file_path)

    print(f"\n原有sheet: {wb.sheetnames}")

    # 1. 在"实验7-档位模拟器"中标注抄袭学生
    if '实验7-档位模拟器' in wb.sheetnames:
        print("\n标注抄袭学生...")
        ws = wb['实验7-档位模拟器']

        # 定义颜色
        plagiarism_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # 红色 - 抄袭
        nosubmit_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')   # 黄色 - 未交
        original_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')  # 绿色 - 原创

        # 找到状态列的索引
        header_row = None
        status_col = None

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value == '状态':
                    status_col = col_idx
                    header_row = row_idx
                    break
            if status_col:
                break

        if status_col:
            print(f"  状态列位置: 第{status_col}列")

            # 遍历数据行，根据状态添加颜色
            colored_count = {'原创': 0, '抄袭': 0, '未交': 0}

            for row_idx in range(header_row + 1, ws.max_row + 1):
                status_cell = ws.cell(row_idx, status_col)
                status = status_cell.value

                if status == '抄袭':
                    # 整行标红
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row_idx, col_idx).fill = plagiarism_fill
                    colored_count['抄袭'] += 1
                elif status == '未交':
                    # 整行标黄
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row_idx, col_idx).fill = nosubmit_fill
                    colored_count['未交'] += 1
                elif status == '原创':
                    # 整行标绿（可选）
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row_idx, col_idx).fill = original_fill
                    colored_count['原创'] += 1

            print(f"  标注完成:")
            print(f"    原创（绿色）: {colored_count['原创']}人")
            print(f"    抄袭（红色）: {colored_count['抄袭']}人")
            print(f"    未交（黄色）: {colored_count['未交']}人")

    # 2. 删除多余的sheet
    sheets_to_delete = []
    sheets_to_keep = ['学期汇总', '实验7-档位模拟器']

    for sheet_name in wb.sheetnames:
        if sheet_name not in sheets_to_keep:
            sheets_to_delete.append(sheet_name)

    if sheets_to_delete:
        print(f"\n删除多余sheet:")
        for sheet_name in sheets_to_delete:
            print(f"  - {sheet_name}")
            del wb[sheet_name]
    else:
        print("\n没有需要删除的sheet")

    print(f"\n最终保留sheet: {wb.sheetnames}")

    # 保存
    wb.save(file_path)
    wb.close()

    print(f"\n文件已保存: {file_path.name}")


def main():
    file_path = Path(__file__).parent.parent / 'docs/teaching/2026-春季/汽服2302B班/docs/汽服2302B班_2026春季学期成绩册.xlsx'

    if not file_path.exists():
        print(f"文件不存在: {file_path}")
        return

    clean_xlsx(file_path)


if __name__ == '__main__':
    main()
