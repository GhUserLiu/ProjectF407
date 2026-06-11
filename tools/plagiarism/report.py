"""
查重报告生成模块
Plagiarism Report Generator

生成详细的查重报告和相似度矩阵
"""

import json
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side,
    GradientFill
)
try:
    from openpyxl.chart import HeatmapChart
    HEATMAP_AVAILABLE = True
except ImportError:
    HeatmapChart = None
    HEATMAP_AVAILABLE = False
from .core import SimilarityResult, SimilarityMethod


@dataclass
class ReportConfig:
    """报告配置"""
    output_dir: Path
    experiment_name: str = '实验报告'
    class_name: str = '未知班级'
    threshold: float = 60.0
    colors: Dict[str, str] = field(default_factory=lambda: {
        'low': '90EE90',      # 浅绿
        'medium': 'FFD700',   # 金色
        'high': 'FFA500',     # 橙色
        'critical': 'FF6347'   # 番茄红
    })


class SimilarityMatrix:
    """相似度矩阵生成器"""

    def __init__(self, student_ids: List[str], names: Dict[str, str]):
        """
        初始化矩阵

        Args:
            student_ids: 学号列表
            names: {学号: 姓名}
        """
        self.student_ids = student_ids
        self.names = names
        self.matrix = {}

        # 初始化矩阵
        for sid1 in student_ids:
            self.matrix[sid1] = {}
            for sid2 in student_ids:
                self.matrix[sid1][sid2] = 0.0

    def set_similarity(self, id1: str, id2: str, value: float):
        """设置相似度值（对称设置）"""
        self.matrix[id1][id2] = value
        self.matrix[id2][id1] = value

    def to_excel(self, ws, start_row: int = 1, start_col: int = 1):
        """
        导出到 Excel

        Args:
            ws: 工作表
            start_row: 起始行
            start_col: 起始列
        """
        # 写入表头
        ws.cell(start_row, start_col).value = '学号\\学号'

        for j, sid in enumerate(self.student_ids):
            name = self.names.get(sid, sid)
            ws.cell(start_row, start_col + j + 1).value = f"{sid}\n{name}"

        for i, sid in enumerate(self.student_ids):
            ws.cell(start_row + i + 1, start_col).value = sid

        # 写入相似度数据并着色
        colors = {
            (0, 40): '90EE90',    # 绿色
            (40, 60): 'FFD700',   # 黄色
            (60, 80): 'FFA500',   # 橙色
            (80, 100): 'FF6347'   # 红色
        }

        for i, sid1 in enumerate(self.student_ids):
            for j, sid2 in enumerate(self.student_ids):
                value = self.matrix[sid1][sid2]
                cell = ws.cell(start_row + i + 1, start_col + j + 1)
                cell.value = round(value, 1) if value > 0 else ''

                # 着色
                for (low, high), color in colors.items():
                    if low <= value < high:
                        cell.fill = PatternFill(
                            start_color=color,
                            end_color=color,
                            fill_type='solid'
                        )
                        break

        # 调整列宽
        for col_idx in range(len(self.student_ids) + 1):
            ws.column_dimensions[start_col + col_idx].width = 12


class PlagiarismReport:
    """查重报告生成器"""

    def __init__(self, config: ReportConfig):
        """
        初始化报告生成器

        Args:
            config: 报告配置
        """
        self.config = config
        self.results: Dict[str, List[SimilarityResult]] = {}
        self.suspicious: List[SimilarityResult] = []
        self.groups: List[Dict] = []

    def add_results(
        self,
        all_results: Dict[str, List[SimilarityResult]],
        suspicious: List[SimilarityResult]
    ):
        """添加检测结果"""
        self.results = all_results
        self.suspicious = suspicious

    def add_groups(self, groups: List[Dict]):
        """添加抄袭团伙信息"""
        self.groups = groups

    def generate_excel(self, filename: str = '查重报告.xlsx'):
        """
        生成 Excel 报告

        Args:
            filename: 输出文件名
        """
        wb = openpyxl.Workbook()

        # 创建各个工作表
        self._create_summary_sheet(wb)
        self._create_detail_sheet(wb)
        self._create_matrix_sheet(wb)
        if self.groups:
            self._create_groups_sheet(wb)
        self._create_statistics_sheet(wb)

        # 保存
        output_path = self.config.output_dir / filename
        wb.save(output_path)
        return output_path

    def _create_summary_sheet(self, wb):
        """创建汇总表"""
        ws = wb.active
        ws.title = "汇总"

        # 标题
        ws['A1'] = f"{self.config.class_name} - {self.config.experiment_name} 查重报告"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')

        ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"可疑阈值: {self.config.threshold}%"

        # 汇总统计
        ws['A5'] = "统计信息"
        ws['A5'].font = Font(bold=True, size=12)

        stats = [
            ("检测人数", len(self.results)),
            ("可疑对数", len(self.suspicious)),
            ("涉嫌抄袭人数", len(self._get_suspicious_students())),
            ("抄袭团伙数", len(self.groups)),
        ]

        for i, (label, value) in enumerate(stats, 6):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value

        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

    def _create_detail_sheet(self, wb):
        """创建详细结果表"""
        ws = wb.create_sheet("详细结果")

        # 表头
        headers = ['学号1', '姓名1', '学号2', '姓名2',
                   '整体相似度', '文本相似度', '代码相似度',
                   '跨组', '共享段落数', '共享代码数']
        ws.append(headers)

        # 样式
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # 数据
        for result in sorted(self.suspicious, key=lambda x: x.overall_similarity, reverse=True):
            metadata = result.metadata
            row = [
                result.student_id,
                metadata.get('name1', ''),
                result.similar_to,
                metadata.get('name2', ''),
                f"{result.overall_similarity:.1f}%",
                f"{result.text_similarity:.1f}%",
                f"{result.code_similarity:.1f}%",
                '是' if result.is_cross_group else '否',
                len(result.shared_paragraphs),
                len(result.shared_code_blocks)
            ]
            ws.append(row)

            # 着色
            last_row = ws.max_row
            if result.overall_similarity >= 85:
                fill_color = self.config.colors['critical']
            elif result.overall_similarity >= 70:
                fill_color = self.config.colors['high']
            else:
                fill_color = self.config.colors['medium']

            for col_idx in range(1, 11):
                ws.cell(last_row, col_idx).fill = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type='solid'
                )

        # 调整列宽
        widths = [12, 12, 12, 12, 12, 12, 12, 8, 12, 12]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

    def _create_matrix_sheet(self, wb):
        """创建相似度矩阵表"""
        # 收集所有学生
        all_students = set(self.results.keys())
        for result in self.suspicious:
            all_students.add(result.student_id)
            all_students.add(result.similar_to)

        student_ids = sorted(list(all_students))
        names = {}

        # 从结果中提取姓名
        for sid, results in self.results.items():
            if results:
                names[sid] = results[0].metadata.get('name1', sid)

        for result in self.suspicious:
            metadata = result.metadata
            names[result.student_id] = metadata.get('name1', result.student_id)
            names[result.similar_to] = metadata.get('name2', result.similar_to)

        # 创建矩阵
        matrix = SimilarityMatrix(student_ids, names)

        # 填充数据
        for sid1, results in self.results.items():
            for result in results:
                matrix.set_similarity(sid1, result.similar_to, result.overall_similarity)

        # 导出到 Excel
        ws = wb.create_sheet("相似度矩阵")
        matrix.to_excel(ws, start_row=2, start_col=2)

        # 添加标题
        ws['A1'] = "相似度矩阵 (%)"
        ws['A1'].font = Font(bold=True, size=14)

    def _create_groups_sheet(self, wb):
        """创建抄袭团伙表"""
        ws = wb.create_sheet("抄袭团伙")

        # 表头
        headers = ['团伙编号', '人数', '成员学号', '成员姓名', '最高相似度']
        ws.append(headers)

        # 样式
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # 数据
        for i, group in enumerate(self.groups, 1):
            members = group['members']
            max_sim = self._get_group_max_similarity(members)

            ws.append([
                i,
                group['size'],
                ', '.join(members),
                '',  # 姓名列留空或填充
                f"{max_sim:.1f}%"
            ])

        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15

    def _create_statistics_sheet(self, wb):
        """创建统计分析表"""
        ws = wb.create_sheet("统计分析")

        # 标题
        ws['A1'] = "相似度分布统计"
        ws['A1'].font = Font(bold=True, size=14)

        # 相似度分布
        ws['A3'] = "相似度区间"
        ws['B3'] = "对数"
        ws['C3'] = "占比"

        for cell in ws[3]:
            cell.font = Font(bold=True)

        # 统计分布
        distributions = [
            (0, 40, '低相似度'),
            (40, 60, '中等相似度'),
            (60, 80, '高相似度'),
            (80, 100, '极高相似度')
        ]

        total = len(self.suspicious) if self.suspicious else 1

        for low, high, label in distributions:
            count = sum(1 for r in self.suspicious if low <= r.overall_similarity < high)
            percentage = (count / total) * 100 if total > 0 else 0

            ws.append([
                f"{label} ({low}-{high}%)",
                count,
                f"{percentage:.1f}%"
            ])

        # 调整列宽
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12

    def _get_suspicious_students(self) -> set:
        """获取所有涉嫌抄袭的学生"""
        students = set()
        for result in self.suspicious:
            students.add(result.student_id)
            students.add(result.similar_to)
        return students

    def _get_group_max_similarity(self, members: List[str]) -> float:
        """获取团伙内最高相似度"""
        max_sim = 0.0

        for sid1 in members:
            for result in self.results.get(sid1, []):
                if result.similar_to in members:
                    max_sim = max(max_sim, result.overall_similarity)

        return max_sim

    def generate_json(self, filename: str = '查重报告.json'):
        """
        生成 JSON 报告

        Args:
            filename: 输出文件名
        """
        output = {
            'meta': {
                'experiment_name': self.config.experiment_name,
                'class_name': self.config.class_name,
                'threshold': self.config.threshold,
                'generated_at': datetime.now().isoformat()
            },
            'summary': {
                'total_students': len(self.results),
                'suspicious_pairs': len(self.suspicious),
                'suspicious_students': len(self._get_suspicious_students()),
                'plagiarism_groups': len(self.groups)
            },
            'suspicious_details': [
                {
                    'student1': r.student_id,
                    'student2': r.similar_to,
                    'overall_similarity': r.overall_similarity,
                    'is_cross_group': r.is_cross_group,
                    'shared_paragraphs': len(r.shared_paragraphs),
                    'shared_code_blocks': len(r.shared_code_blocks)
                }
                for r in sorted(self.suspicious, key=lambda x: x.overall_similarity, reverse=True)
            ],
            'groups': self.groups
        }

        output_path = self.config.output_dir / filename
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(output, f, ensure_ascii=False, indent=2)

        return output_path

    def generate_html(self, filename: str = '查重报告.html'):
        """
        生成 HTML 报告

        Args:
            filename: 输出文件名
        """
        html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{self.config.experiment_name} 查重报告</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        h1 {{ color: #333; }}
        h2 {{ color: #666; border-bottom: 1px solid #ccc; padding-bottom: 10px; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #4472C4; color: white; }}
        .critical {{ background-color: #FF6347; color: white; }}
        .high {{ background-color: #FFA500; }}
        .medium {{ background-color: #FFD700; }}
        .stats {{ display: flex; gap: 20px; margin: 20px 0; }}
        .stat-box {{ border: 1px solid #ddd; padding: 15px; border-radius: 5px; }}
        .stat-number {{ font-size: 24px; font-weight: bold; color: #4472C4; }}
    </style>
</head>
<body>
    <h1>{self.config.class_name} - {self.config.experiment_name} 查重报告</h1>
    <p>生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>

    <div class="stats">
        <div class="stat-box">
            <div>检测人数</div>
            <div class="stat-number">{len(self.results)}</div>
        </div>
        <div class="stat-box">
            <div>可疑对数</div>
            <div class="stat-number">{len(self.suspicious)}</div>
        </div>
        <div class="stat-box">
            <div>涉嫌抄袭人数</div>
            <div class="stat-number">{len(self._get_suspicious_students())}</div>
        </div>
    </div>

    <h2>高相似度对详情 (≥{self.config.threshold}%)</h2>
    <table>
        <tr>
            <th>学号1</th>
            <th>姓名1</th>
            <th>学号2</th>
            <th>姓名2</th>
            <th>相似度</th>
            <th>跨组</th>
        </tr>
"""

        for result in sorted(self.suspicious, key=lambda x: x.overall_similarity, reverse=True):
            metadata = result.metadata
            css_class = 'critical' if result.overall_similarity >= 85 else 'high' if result.overall_similarity >= 70 else 'medium'

            html_content += f"""
        <tr class="{css_class}">
            <td>{result.student_id}</td>
            <td>{metadata.get('name1', '')}</td>
            <td>{result.similar_to}</td>
            <td>{metadata.get('name2', '')}</td>
            <td>{result.overall_similarity:.1f}%</td>
            <td>{'是' if result.is_cross_group else '否'}</td>
        </tr>
"""

        html_content += """
    </table>

    <h2>抄袭团伙</h2>
    <table>
        <tr>
            <th>团伙编号</th>
            <th>人数</th>
            <th>成员</th>
        </tr>
"""

        for i, group in enumerate(self.groups, 1):
            members = ', '.join(group['members'])
            html_content += f"""
        <tr>
            <td>{i}</td>
            <td>{group['size']}</td>
            <td>{members}</td>
        </tr>
"""

        html_content += """
    </table>
</body>
</html>
"""

        output_path = self.config.output_dir / filename
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        return output_path
