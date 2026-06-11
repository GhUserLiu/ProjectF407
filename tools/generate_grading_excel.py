#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生成实验成绩表（Excel格式）
Generate Experiment Grading Sheet (Excel Format)
"""

import json
import argparse
from pathlib import Path
from datetime import datetime
from typing import Dict, List

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    exit(1)


def generate_grading_sheet(
    data: List[Dict],
    output_path: Path,
    experiment_name: str = "实验报告",
    class_name: str = "未知班级"
) -> Path:
    """
    生成Excel成绩表

    Args:
        data: 评分结果数据
        output_path: 输出文件路径
        experiment_name: 实验名称
        class_name: 班级名称

    Returns:
        输出文件路径
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "成绩表"

    # 样式定义
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=side_style, right=side_style,
        top=side_style, bottom=side_style
    )

    # 标题
    title = f"{class_name} - {experiment_name} 成绩表"
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:K1')

    ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(size=10, color='808080')
    ws.merge_cells('A2:K2')

    # 表头
    headers = ['学号', '姓名', '总分', '等级', '团队协作', '实验态度',
                '原理认知', '实验完成度', '代码质量', '报告质量', '备注']
    ws.append([''] * len(headers))  # 空行
    ws.append(headers)

    # 设置表头样式
    header_row = ws.max_row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # 按学号排序
    sorted_data = sorted(data, key=lambda x: x['student_id'])

    # 数据行
    for item in sorted_data:
        category_scores = item.get('category_scores', {})

        # 提取各类别分数
        team = category_scores.get('team_collaboration', {})
        attitude = category_scores.get('attitude', {})
        principle = category_scores.get('principle_understanding', {})
        completion = category_scores.get('completion', {})
        code = category_scores.get('code_quality', {})
        report = category_scores.get('report_quality', {})

        row = [
            item['student_id'],
            item['name'],
            item['total_score'],
            item['grade'],
            f"{team.get('earned', 0):.1f}",
            f"{attitude.get('earned', 0):.1f}",
            f"{principle.get('earned', 0):.1f}",
            f"{completion.get('earned', 0):.1f}",
            f"{code.get('earned', 0):.1f}",
            f"{report.get('earned', 0):.1f}",
            ''
        ]
        ws.append(row)

        # 设置数据行样式
        data_row = ws.max_row
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=data_row, column=col_num)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # 等列着色
            if col_num == 4:  # 等级列
                grade = item['grade']
                if grade == 'A':
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif grade == 'B':
                    cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                elif grade == 'C':
                    cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                elif grade == 'D':
                    cell.fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
                elif grade == 'F':
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    # 设置列宽
    column_widths = {
        'A': 15,  # 学号
        'B': 12,  # 姓名
        'C': 8,   # 总分
        'D': 8,   # 等级
        'E': 10,  # 团队协作
        'F': 10,  # 实验态度
        'G': 10,  # 原理认知
        'H': 10,  # 实验完成度
        'I': 10,  # 代码质量
        'J': 10,  # 报告质量
        'K': 20   # 备注
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 设置行高
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[header_row].height = 20

    # 保存文件
    wb.save(output_path)
    return output_path


def generate_detail_sheet(
    data: List[Dict],
    output_path: Path,
    experiment_name: str = "实验报告",
    class_name: str = "未知班级"
) -> Path:
    """
    生成详细评分表（包含反馈）

    Args:
        data: 评分结果数据
        output_path: 输出文件路径
        experiment_name: 实验名称
        class_name: 班级名称

    Returns:
        输出文件路径
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "详细评分"

    # 样式定义
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    border = Border(
        left=side_style, right=side_style,
        top=side_style, bottom=side_style
    )

    # 按学号排序
    sorted_data = sorted(data, key=lambda x: x['student_id'])

    # 表头
    headers = ['学号', '姓名', '总分', '等级', '类别', '得分', '满分', '反馈']
    ws.append(headers)

    # 设置表头样式
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # 为每个学生生成多行（每个评分类别一行）
    for item in sorted_data:
        student_id = item['student_id']
        name = item['name']
        total_score = item['total_score']
        grade = item['grade']
        category_scores = item.get('category_scores', {})

        # 第一行：学生基本信息
        first_row_num = ws.max_row + 1
        ws.cell(row=first_row_num, column=1, value=student_id)
        ws.cell(row=first_row_num, column=2, value=name)
        ws.cell(row=first_row_num, column=3, value=total_score)
        ws.cell(row=first_row_num, column=4, value=grade)

        # 等级单元格着色
        grade_cell = ws.cell(row=first_row_num, column=4)
        if grade == 'A':
            grade_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif grade == 'B':
            grade_cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        elif grade == 'C':
            grade_cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        elif grade == 'D':
            grade_cell.fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
        elif grade == 'F':
            grade_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # 各类别评分行
        for cat_id, cat_score in category_scores.items():
            row_num = ws.max_row + 1
            ws.cell(row=row_num, column=5, value=cat_score['name'])
            ws.cell(row=row_num, column=6, value=cat_score['earned'])
            ws.cell(row=row_num, column=7, value=cat_score['possible'])

            # 反馈（合并所有反馈项）
            feedback_list = cat_score.get('feedback', [])
            feedback_text = '; '.join(str(f) for f in feedback_list)
            ws.cell(row=row_num, column=8, value=feedback_text)

        # 添加空行分隔
        ws.append([''] * 8)

    # 设置列宽
    column_widths = {
        'A': 15,  # 学号
        'B': 12,  # 姓名
        'C': 8,   # 总分
        'D': 8,   # 等级
        'E': 15,  # 类别
        'F': 8,   # 得分
        'G': 8,   # 满分
        'H': 50   # 反馈
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 设置列自动换行
    for row in ws.iter_rows(min_row=2, max_col=8):
        row[7].alignment = Alignment(wrap_text=True, vertical='top')

    # 保存文件
    wb.save(output_path)
    return output_path


# 用于边框样式
side_style = openpyxl.styles.borders.Side(style='thin')


def main():
    parser = argparse.ArgumentParser(
        description='生成实验成绩表（Excel格式）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例用法:
  # 生成成绩表
  python tools/generate_grading_excel.py

  # 指定输入文件
  python tools/generate_grading_excel.py --input results/grading_results.json

  # 生成详细评分表
  python tools/generate_grading_excel.py --detailed
        """
    )

    parser.add_argument(
        '--input',
        type=Path,
        default=Path('docs/teaching/2026-春季/汽服2302B班/07-car-gear/results/grading_results.json'),
        help='评分结果JSON文件路径'
    )

    parser.add_argument(
        '--output-dir',
        type=Path,
        default=Path('docs/teaching/2026-春季/汽服2302B班/07-car-gear/results'),
        help='输出目录路径'
    )

    parser.add_argument(
        '--experiment-name',
        type=str,
        default='档位实验',
        help='实验名称'
    )

    parser.add_argument(
        '--class-name',
        type=str,
        default='汽服2302B班',
        help='班级名称'
    )

    parser.add_argument(
        '--detailed',
        action='store_true',
        help='生成详细评分表（包含反馈）'
    )

    parser.add_argument(
        '--both',
        action='store_true',
        help='同时生成成绩表和详细评分表'
    )

    args = parser.parse_args()

    # 加载数据
    if not args.input.exists():
        print(f"错误: 输入文件不存在: {args.input}")
        return 1

    with open(args.input, 'r', encoding='utf-8') as f:
        data = json.load(f)

    print(f"加载了 {len(data)} 条评分记录")

    # 创建输出目录
    args.output_dir.mkdir(parents=True, exist_ok=True)

    # 生成文件
    if args.detailed or args.both:
        output_path = args.output_dir / f"{args.class_name}_详细评分表.xlsx"
        generate_detail_sheet(data, output_path, args.experiment_name, args.class_name)
        print(f"详细评分表: {output_path}")

    if not args.detailed or args.both:
        output_path = args.output_dir / f"{args.class_name}_{args.experiment_name}_成绩表.xlsx"
        generate_grading_sheet(data, output_path, args.experiment_name, args.class_name)
        print(f"成绩表: {output_path}")

    return 0


if __name__ == '__main__':
    exit(main())
