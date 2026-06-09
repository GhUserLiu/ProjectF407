#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
学期成绩册生成系统
- 提取并清理学生姓名
- 合并多次实验成绩
- 生成学期汇总表
"""

import json
import re
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed")
    print("Please install: pip install openpyxl")
    sys.exit(1)

# ==================== 配置 ====================

# 提交时间（学号 -> 时间字符串）
SUBMISSION_TIMES = {
    '23071140226': '2026-06-04 16:30:28', '23071140231': '2026-06-04 16:34:35',
    '23071140230': '2026-06-06 02:28:52', '23071140217': '2026-06-06 11:04:24',
    '23071140221': '2026-06-06 12:15:50', '23071140210': '2026-06-06 12:20:02',
    '23071140233': '2026-06-06 17:00:24', '23071140235': '2026-06-06 21:41:16',
    '23071140234': '2026-06-06 21:57:37', '23071140240': '2026-06-06 22:42:25',
    '23071140241': '2026-06-06 22:48:54', '23071140238': '2026-06-06 22:56:51',
    '23071140239': '2026-06-06 23:04:58', '23071140227': '2026-06-06 23:09:14',
    '23071140201': '2026-06-07 01:05:53', '23071140229': '2026-06-07 08:33:29',
    '23071140204': '2026-06-07 08:50:07', '23071140208': '2026-06-07 08:54:01',
    '23071140237': '2026-06-07 12:21:56', '23071140213': '2026-06-07 14:09:43',
    '23071140219': '2026-06-07 14:11:53', '23071140202': '2026-06-07 14:23:46',
    '23071140214': '2026-06-07 19:20:27', '23071140228': '2026-06-07 19:36:11',
    '23071140222': '2026-06-07 19:40:57', '23071140232': '2026-06-07 19:58:16',
    '23071140215': '2026-06-07 20:45:13', '23071140206': '2026-06-07 20:53:23',
    '23071140224': '2026-06-07 21:03:28', '23071140211': '2026-06-07 22:41:34',
    '23071140220': '2026-06-07 22:52:21', '23071140203': '2026-06-07 23:27:32',
    '23071140216': '2026-06-08 09:46:04', '23071140223': '2026-06-08 09:58:02',
    '23071140236': '2026-06-08 10:15:27'
}

# 抄弊学生列表
PLAGIARISM_STUDENTS = {
    '23071140217', '23071140216', '23071140214', '23071140228',
    '23071140233', '23071140220', '23071140223', '23071140213',
    '23071140219', '23071140204', '23071140208'
}

# ==================== 姓名清理函数 ====================

def clean_name(name):
    """清理学生姓名，去掉组别前缀和多余信息"""
    if not name:
        return ''

    # 去掉常见的组别前缀
    prefixes = ['第1组-', '第2组-', '第3组-', '第4组-', '第5组-',
                '第6组-', '第7组-', '第8组-', '第9组-', '第十组-',
                '汽服2302B班-']

    for prefix in prefixes:
        if name.startswith(prefix):
            name = name[len(prefix):]
            break

    # 去掉文件扩展名
    name = name.replace('.docx', '').replace('.doc', '')

    # 去掉常见的模板后缀
    suffixes = ['实验报告模板', '实验报告模板(1)', '实验报告模板（',
                '汽车挡位模拟设计', '设计']
    for suffix in suffixes:
        if suffix in name:
            name = name.replace(suffix, '')

    # 如果姓名过长或包含异常字符，返回空
    if len(name) > 10 or '实验报告' in name or '心得体会' in name:
        return ''

    return name.strip()

def load_student_names(filepath):
    """加载学生姓名数据"""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)

        name_map = {}
        for student in data:
            sid = student['id']
            raw_name = student.get('name', '')
            clean = clean_name(raw_name)
            if clean:
                name_map[sid] = clean

        return name_map
    except Exception as e:
        print(f"Error loading student names: {e}")
        return {}

def load_student_info_from_zip(filepath):
    """从zip文件中提取的学生信息"""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Warning: Cannot load student_info_from_zip.json: {e}")
        return {}

def load_evaluation_data(filepath):
    """加载评估数据"""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading evaluation data: {e}")
        return []

def calculate_score(evaluation_data, is_plagiarism):
    """计算单项得分"""
    if is_plagiarism:
        return 0, {}

    scores = evaluation_data.get('scores', {})
    detail_scores = {
        '团队协作': min(scores.get('team_collaboration', 0), 5),
        '实验原理': min(scores.get('principle_understanding', 0), 10),
        '实验完成度': min(scores.get('completion', 0), 35),
        '代码质量': min(scores.get('code_quality', 0), 30),
        '报告质量': min(scores.get('report_quality', 0), 10),
        '实验态度': min(scores.get('attitude', 0), 10)
    }

    total = sum(detail_scores.values())
    return total, detail_scores

def get_grade(score):
    """根据分数获取等级"""
    if score >= 90:
        return 'A', '优秀'
    elif score >= 80:
        return 'B', '良好'
    elif score >= 70:
        return 'C', '中等'
    elif score >= 60:
        return 'D', '及格'
    else:
        return 'F', '不及格'

# ==================== 主程序 ====================

def main():
    print("=" * 60)
    print("汽服2302B班 - 2026春季学期 成绩册生成系统")
    print("=" * 60)

    # 加载数据
    scripts_dir = '.'
    students_file = f'{scripts_dir}/data/students.json'
    eval_file = f'{scripts_dir}/data/evaluations.json'
    zip_info_file = f'{scripts_dir}/../07-car-gear/scripts/data/student_info_from_zip.json'

    name_map = load_student_names(students_file)
    evaluations = load_evaluation_data(eval_file)
    zip_info = load_student_info_from_zip(zip_info_file)

    print(f"\n数据加载完成:")
    print(f"  学生姓名(students.json): {len(name_map)} 人")
    print(f"  评估记录: {len(evaluations)} 条")
    print(f"  Zip文件信息: {len(zip_info)} 人")

    # 使用zip信息作为主要姓名来源
    if zip_info:
        name_map = {sid: info['name'] for sid, info in zip_info.items()}
        print(f"  使用Zip文件中的姓名: {len(name_map)} 人")

    # 创建评估映射
    eval_map = {s['student_id']: s for s in evaluations}

    # 生成第七次实验成绩
    exp7_results = []
    for sid, time_str in SUBMISSION_TIMES.items():
        is_plagiarism = sid in PLAGIARISM_STUDENTS
        eval_data = eval_map.get(sid, {})

        # 获取姓名和实验报告状态
        name = name_map.get(sid, '')
        has_report = zip_info.get(sid, {}).get('has_report', True) if zip_info else True

        if is_plagiarism:
            score = 0
            status = "抄袭"
            details = {}
        elif not has_report:
            score = 0
            status = "未提交报告"
            details = {}
        else:
            score, details = calculate_score(eval_data, False)
            status = "原创" if score > 0 else "未完成"

        grade, grade_label = get_grade(score)

        exp7_results.append({
            '学号': sid,
            '姓名': name,
            '提交时间': time_str,
            '总分': score,
            '等级': grade,
            '等级标签': grade_label,
            '状态': status,
            '有实验报告': has_report,
            '团队协作': details.get('团队协作', 0),
            '实验原理': details.get('实验原理', 0),
            '实验完成度': details.get('实验完成度', 0),
            '代码质量': details.get('代码质量', 0),
            '报告质量': details.get('报告质量', 0),
            '实验态度': details.get('实验态度', 0)
        })

    # 按提交时间排序
    exp7_results.sort(key=lambda x: x['提交时间'])

    # 统计信息
    total = len(exp7_results)
    plagiarism = sum(1 for r in exp7_results if r['状态'] == '抄袭')
    no_report = sum(1 for r in exp7_results if r['状态'] == '未提交报告')
    incomplete = sum(1 for r in exp7_results if r['状态'] == '未完成')
    original = total - plagiarism - no_report - incomplete
    avg_score = sum(r['总分'] for r in exp7_results if r['总分'] > 0) / max(original, 1)

    print(f"\n第七次实验统计:")
    print(f"  总人数: {total}")
    print(f"  抄袭: {plagiarism} 人")
    print(f"  未提交报告: {no_report} 人")
    print(f"  未完成: {incomplete} 人")
    print(f"  原创完成: {original} 人")
    print(f"  平均分: {avg_score:.1f}")

    # 生成Excel
    generate_semester_excel(exp7_results, {
        'total': total,
        'plagiarism': plagiarism,
        'no_report': no_report,
        'incomplete': incomplete,
        'original': original,
        'avg': avg_score
    })

    print("\n成绩册生成完成！")
    print("文件: 汽服2302B班_2026春季学期成绩册.xlsx")

def generate_semester_excel(exp7_data, stats):
    """生成学期成绩册Excel"""

    wb = openpyxl.Workbook()

    # ==================== Sheet 1: 学期汇总 ====================
    ws_summary = wb.active
    ws_summary.title = "学期汇总"

    # 标题
    ws_summary['A1'] = '汽服2302B班 - 2026春季学期 实验成绩汇总'
    ws_summary.merge_cells('A1:F1')
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary['A1'].alignment = Alignment(horizontal='center')

    # 表头
    headers = ['学号', '姓名', '实验1', '实验2', '实验3', '实验4',
               '实验5', '实验6', '实验7', '平均分', '总评']
    row = 3
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # 数据行
    for result in exp7_data:
        row += 1
        ws_summary.cell(row, 1, result['学号'])
        ws_summary.cell(row, 2, result['姓名'])
        ws_summary.cell(row, 3, '')  # 实验1-6 暂无数据
        ws_summary.cell(row, 4, '')
        ws_summary.cell(row, 5, '')
        ws_summary.cell(row, 6, '')
        ws_summary.cell(row, 7, '')
        ws_summary.cell(row, 8, '')
        ws_summary.cell(row, 9, result['总分'])  # 实验7
        ws_summary.cell(row, 10, result['总分'])  # 平均分（目前只有一次实验）

        # 总评等级
        grade_cell = ws_summary.cell(row, 11, result['等级'])
        if result['状态'] == '抄袭':
            grade_cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        elif result['状态'] == '未提交报告':
            grade_cell.fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
        elif result['总分'] >= 90:
            grade_cell.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')

    # 设置列宽
    for col in range(1, 12):
        ws_summary.column_dimensions[get_column_letter(col)].width = 12

    # ==================== Sheet 2: 实验7详细成绩 ====================
    ws_exp7 = wb.create_sheet('实验7-档位模拟器')

    # 标题
    ws_exp7['A1'] = '第七次实验：汽车档位模拟器设计'
    ws_exp7.merge_cells('A1:M1')
    ws_exp7['A1'].font = Font(size=14, bold=True)
    ws_exp7['A1'].alignment = Alignment(horizontal='center')

    # 统计信息
    ws_exp7['A2'] = f'统计: 总{stats["total"]}人 | 抄袭{stats["plagiarism"]}人 | 未提交报告{stats["no_report"]}人 | 未完成{stats["incomplete"]}人 | 原创{stats["original"]}人 | 平均{stats["avg"]:.1f}分'
    ws_exp7.merge_cells('A2:M2')
    ws_exp7['A2'].font = Font(size=10, italic=True)

    # 表头
    headers = ['排名', '学号', '姓名', '提交时间', '总分', '等级', '状态',
               '团队协作', '实验原理', '实验完成度', '代码质量', '报告质量', '实验态度']
    row = 4
    for col, header in enumerate(headers, 1):
        cell = ws_exp7.cell(row, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # 数据行
    rank = 1
    for result in exp7_data:
        row += 1
        ws_exp7.cell(row, 1, rank if result['总分'] > 0 else '-')
        ws_exp7.cell(row, 2, result['学号'])
        ws_exp7.cell(row, 3, result['姓名'])
        ws_exp7.cell(row, 4, result['提交时间'])
        ws_exp7.cell(row, 5, result['总分'])
        ws_exp7.cell(row, 6, result['等级'])
        ws_exp7.cell(row, 7, result['状态'])
        ws_exp7.cell(row, 8, result['团队协作'])
        ws_exp7.cell(row, 9, result['实验原理'])
        ws_exp7.cell(row, 10, result['实验完成度'])
        ws_exp7.cell(row, 11, result['代码质量'])
        ws_exp7.cell(row, 12, result['报告质量'])
        ws_exp7.cell(row, 13, result['实验态度'])

        # 根据状态设置填充色
        if result['状态'] == '抄袭':
            fill_color = 'FFCCCC'
        elif result['状态'] == '未完成':
            fill_color = 'FFFFCC'
        elif result['总分'] >= 90:
            fill_color = 'CCFFCC'
        else:
            fill_color = None

        if fill_color:
            for col in range(1, 14):
                ws_exp7.cell(row, col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        if result['总分'] > 0:
            rank += 1

    # 设置列宽
    column_widths = [6, 12, 12, 18, 6, 6, 8, 10, 10, 12, 10, 10, 10]
    for col, width in enumerate(column_widths, 1):
        ws_exp7.column_dimensions[get_column_letter(col)].width = width

    # 保存文件
    output_file = '../docs/汽服2302B班_2026春季学期成绩册.xlsx'
    wb.save(output_file)
    print(f"已保存: {output_file}")

if __name__ == '__main__':
    main()
